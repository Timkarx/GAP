from openpyxl import load_workbook

match_sheet = load_workbook(filename="sample_prem.xlsx").active

#This function takes returns a list of matches,
#with each match being a nested list

def match_list(worksheet, start_col, end_col):
    """This function returns information about matches"""
    matches_played = []
    for match in worksheet.iter_rows(min_col = start_col,
                                    max_col = end_col,
                                     values_only=True):
        matches_played.append(match)
    del matches_played[0]
    return(matches_played)
