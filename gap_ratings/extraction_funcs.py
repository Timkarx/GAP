from openpyxl import load_workbook


def team_list(worksheet, column):
    """This function returns an ordered list of all the teams in the list"""
    
    teams = []
    for team in worksheet[column]:
        teams.append(team.value)

#Delete duplicates and HomeTeam and sort
    if "HomeTeam" in teams:
        teams.remove("HomeTeam")
    elif "AwayTeam" in teams:
        teams.remove("AwayTeam")
    elif "Referee" in teams:
        teams.remove("Referee")
    
    sor_teams = sorted(set(teams))
    return (sor_teams)


def match_list(worksheet, start_col, end_col):
    """This function returns a list with information about matches"""
    
    matches_played = []
    for match in worksheet.iter_rows(min_col = start_col,
                                    max_col = end_col,
                                     values_only=True):
        matches_played.append(match)
    del matches_played[0]
    return(matches_played)


def load_sheet(worksheet_path):
    """Use rstrings when writing the path"""
    """This function loads the excel spreadsheet using openpyxl"""
    
    return (load_workbook(filename = worksheet_path).active)


def load_league(league,start_year,end_year):
    """Return a list of lists of games for each season"""
    
    list = []
    for season_start,season_end in zip(range(start_year,end_year),
                                       range(start_year+1,end_year+1)):
        season_sheet = load_sheet(fr'C:\Users\Tim\Desktop\Programming\Python' +
                                  fr'\Football_Team_Performance_Calculator\foot_data' +
                                  fr'\{league}\epl_{season_start}' +
                                  fr'{season_end}.xlsx')
        list.append(season_sheet)
    return(list)


def extract_league_data(team_column, season_list):
    """This function discards all the useless data in the spreadsheet"""

    season_teams_list = []
    season_match_list = []
    for season_sheet in season_list:
        season_teams_list.append(team_list(season_sheet,team_column))
        if season_list.index(season_sheet)<3:
            matches_played = match_list(season_sheet,3,6)
            season_match_list.append(matches_played)
        else:
            matches_played = match_list(season_sheet,4,7)
            season_match_list.append(matches_played)
    return(season_teams_list, season_match_list)

