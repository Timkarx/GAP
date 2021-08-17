from openpyxl import load_workbook


def team_list(worksheet, column):
    """Returns an ordered list of teams"""

    teams = []
    for team in worksheet[column]:
        teams.append(team.value)

    if "HomeTeam" in teams:
        teams.remove("HomeTeam")
    elif "AwayTeam" in teams:
        teams.remove("AwayTeam")
    elif "Referee" in teams:
        teams.remove("Referee")

    return sorted(set(teams))


def match_list(worksheet, start_col, end_col):
    """Returns a list with info about matches"""

    matches_played = []
    for match in worksheet.iter_rows(min_col=start_col,
                                     max_col=end_col,
                                     values_only=True):
        match_l = list(match)
        matches_played.append(match_l)
    del matches_played[0]
    return matches_played


def load_sheet(worksheet_path):
    """Use rstrings when writing the path"""
    """Loads the excel spreadsheet using openpyxl"""

    return load_workbook(filename=worksheet_path).active


def load_leagues(league, start_year, end_year):
    """Return a list of lists of games for each season"""

    league_list = []
    for season_start, season_end in zip(range(start_year, end_year),
                                        range((start_year + 1), (end_year + 1))):
        season_sheet = load_sheet(fr'C:\Programming\Python' +
                                  fr'\Football_Team_Performance_Calculator\foot_data\input' +
                                  fr'\{league}\{league}_{season_start}' +
                                  fr'{season_end}.xlsx')
        league_list.append(season_sheet)
    return league_list


def extract_league_data(team_column, season_list, first_col, last_col):
    """
    Discards all the useless data in the spreadsheet
    Note: team_column argument indicates which column the team names are located in
    Note2: each element of season_list is a list of games per season

    """

    season_teams_list = []
    season_match_list = []
    for season_sheet in season_list:
        season_teams_list.append(team_list(season_sheet, team_column))
        season_match_list.append(match_list(season_sheet, first_col, last_col))
    return season_teams_list, season_match_list
