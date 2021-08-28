"""This file edits excel spreadsheets. It sums two columns of stats and stores the result in a new column"""
import sys
sys.path.insert(1,r'C:\Users\Tim\AppData\Local\Packages\PythonSoftwareFoundation.Python.3.9_qbz5n2kfra8p0\LocalCache\local-packages\Python39\site-packages')

from openpyxl import *


leagues = ["Premier_League","La_Liga","Serie_A","Ligue_1","Bundesliga"]

for league in leagues:
    for season_start, season_end in zip(range(9, 21),
                                            range(10, 22)):

        workbook = load_workbook(filename=fr'C:\Programming\Python' +
                                    fr'\Football_Team_Performance_Calculator\foot_data\input' +
                                    fr'\{league}\{league}_{season_start}' +
                                     fr'{season_end}.xlsx')


        season_sheet = workbook.active
        season_sheet.insert_cols(23,4)

        for m in range(1,500):

            if season_sheet.cell(row=m, column=11).value == None:
                continue
            else:
                home_shots = season_sheet.cell(row=m, column=11).value
                home_shotsTar = season_sheet.cell(row=m, column=13).value
                home_corners = season_sheet.cell(row=m, column=17).value
                away_shots = season_sheet.cell(row=m, column=12).value
                away_shots_Tar = season_sheet.cell(row=m, column=14).value
                away_corners = season_sheet.cell(row=m, column=18).value

                #Create a column with Shots + Corners
                season_sheet.cell(row=m,column=23).value = home_shots + home_corners
                season_sheet.cell(row=m, column=24).value = away_shots + away_corners

                #Create a column with Shots on Target + Corners
                season_sheet.cell(row=m, column=25).value = home_shotsTar + home_corners
                season_sheet.cell(row=m, column=26).value = away_shots_Tar + away_corners


        workbook.save(filename=fr'C:\Programming\Python' +
                                            fr'\Football_Team_Performance_Calculator\foot_data\input' +
                                            fr'\{league}\{league}_{season_start}' +
                                            fr'{season_end}.xlsx')
